"""
文件导入处理模块
支持从 Excel(.xlsx/.xls) 导入题库
适配通信网络班题库格式：
- 单选 Sheet：序号 | 题干 | 可选项A | 可选项B | 可选项C | 可选项D | 正确答案
- 多选 Sheet：序号 | 题干 | 选项A~F | 正确答案
- 判断 Sheet：序号 | 题干 | 选项A(正确) | 选项B(错误) | 答案(A/B)
"""
import os
import uuid
from openpyxl import load_workbook


class ImportHandler:
    """题库导入处理器"""

    def import_from_excel(self, filepath):
        """从 Excel 文件导入题目，自动识别多个 Sheet"""
        wb = load_workbook(filepath, data_only=True)
        all_questions = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            questions = self._parse_sheet(ws, sheet_name)
            all_questions.extend(questions)

        wb.close()
        return all_questions

    def _parse_sheet(self, ws, sheet_name):
        """解析单个 Sheet"""
        questions = []
        sheet_lower = sheet_name.lower().strip()

        # 判断题型
        if '多选' in sheet_lower:
            q_type = 'multiple'
        elif '判断' in sheet_lower or '是非' in sheet_lower:
            q_type = 'judge'
        else:
            q_type = 'single'  # 默认单选题

        # 获取表头
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # 构建列映射
        col_map = self._build_column_map(headers, q_type)

        # 解析数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue

            try:
                question = self._parse_row(row, col_map, q_type)
                if question and question.get('content'):
                    question['id'] = str(uuid.uuid4())[:8]
                    questions.append(question)
            except Exception as e:
                print(f"Sheet[{sheet_name}] 第{row_idx}行解析失败: {e}")
                continue

        return questions

    def _build_column_map(self, headers, q_type):
        """根据表头构建列名到索引的映射"""
        col_map = {}
        header_lower = [h.lower().strip() for h in headers]

        # 题干列 - 多种可能名称
        content_keywords = ['题干', '题目内容', '题目', '问题', 'question', 'content', '题干（题目内容）']
        # 正确答案列
        answer_keywords = ['正确答案', '答案', 'answer']
        # 选项列
        option_keywords = ['可选项a', '可选项b', '可选项c', '可选项d', '可选项e', '可选项f',
                           '选项a', '选项b', '选项c', '选项d', '选项e', '选项f']

        for idx, h in enumerate(header_lower):
            # 匹配题干
            for kw in content_keywords:
                if kw in h:
                    col_map['content'] = idx
                    break
            # 匹配答案
            for kw in answer_keywords:
                if kw == h:
                    col_map['answer'] = idx
                    break
            # 匹配选项
            for opt_kw in option_keywords:
                if h == opt_kw:
                    key = opt_kw[-1].upper()  # 提取 A, B, C...
                    if 'options' not in col_map:
                        col_map['options'] = {}
                    col_map['options'][key] = idx
                    break

        return col_map

    def _parse_row(self, row, col_map, q_type):
        """解析一行数据为题目"""
        def get_val(idx):
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ''

        content = get_val(col_map.get('content'))
        if not content:
            return None

        answer = get_val(col_map.get('answer')).upper()

        question = {
            'type': q_type,
            'content': content,
            'answer': answer,
            'explanation': ''
        }

        # 解析选项
        if q_type == 'single':
            # 单选题：选项 A~D
            options = []
            opt_map = col_map.get('options', {})
            for key in ['A', 'B', 'C', 'D']:
                idx = opt_map.get(key)
                if idx is not None:
                    opt_text = get_val(idx)
                    if opt_text:
                        options.append({'key': key, 'text': opt_text})
            question['options'] = options

        elif q_type == 'multiple':
            # 多选题：选项 A~F
            options = []
            opt_map = col_map.get('options', {})
            for key in ['A', 'B', 'C', 'D', 'E', 'F']:
                idx = opt_map.get(key)
                if idx is not None:
                    opt_text = get_val(idx)
                    if opt_text:
                        options.append({'key': key, 'text': opt_text})
            question['options'] = options

        elif q_type == 'judge':
            # 判断题：选项 A=正确, B=错误
            opt_map = col_map.get('options', {})
            opt_a_idx = opt_map.get('A')
            opt_b_idx = opt_map.get('B')
            opt_a = get_val(opt_a_idx) if opt_a_idx is not None else '正确'
            opt_b = get_val(opt_b_idx) if opt_b_idx is not None else '错误'
            question['options'] = [
                {'key': 'A', 'text': opt_a},
                {'key': 'B', 'text': opt_b}
            ]

        return question
